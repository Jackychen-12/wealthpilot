"""持仓导入路由 — CSV/Excel 批量导入 + 截图 OCR 识别。"""

import base64
import csv
import io
import json
from datetime import date

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlmodel import Session

from wealthpilot.models.portfolio import PortfolioHolding
from wealthpilot.settings import get_settings
from wealthpilot.storage.db import get_session

router = APIRouter(prefix="/portfolio", tags=["import"])


@router.post("/import/csv")
async def import_csv(file: UploadFile = File(...), db: Session = Depends(get_session)):
    """从 CSV/Excel 批量导入持仓。

    支持格式（任意列名包含关键字即可匹配）：
    - 基金代码, 基金名称, 持有份额, 成本净值, 买入日期, 类型
    - fund_code, fund_name, shares, cost_price, buy_date, category

    也支持天天基金导出的格式。
    """
    if not file.filename:
        raise HTTPException(400, "未提供文件")

    content = await file.read()
    ext = file.filename.rsplit(".", 1)[-1].lower()

    rows: list[dict] = []

    if ext == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v or "") for v in row])))
        except Exception as e:
            raise HTTPException(400, f"Excel 解析失败: {e}") from e
    else:
        raise HTTPException(400, f"不支持的文件格式: {ext}，请上传 .csv 或 .xlsx")

    if not rows:
        raise HTTPException(400, "文件内容为空")

    # 智能列匹配
    imported = []
    for row in rows:
        fund_code = _find_value(row, ["基金代码", "fund_code", "代码", "编码"])
        fund_name = _find_value(row, ["基金名称", "fund_name", "名称", "基金简称"])
        shares = _find_float(row, ["持有份额", "shares", "份额", "持有", "数量"])
        cost_price = _find_float(row, ["成本净值", "cost_price", "成本", "买入净值", "成本价"])
        buy_date_str = _find_value(row, ["买入日期", "buy_date", "日期", "购买日期"])
        category = _find_value(row, ["类型", "category", "基金类型", "分类"]) or "equity"

        if not fund_code or not shares:
            continue

        buy_date_val = _parse_date(buy_date_str) if buy_date_str else date.today()

        holding = PortfolioHolding(
            fund_code=fund_code.strip(),
            fund_name=fund_name or fund_code,
            shares=shares,
            cost_price=cost_price or 1.0,
            buy_date=buy_date_val,
            category=_normalize_category(category),
        )
        db.add(holding)
        imported.append({"fund_code": fund_code, "fund_name": fund_name, "shares": shares})

    db.commit()
    return {"status": "ok", "imported_count": len(imported), "items": imported}


@router.post("/import/ocr")
async def import_ocr(file: UploadFile = File(...), db: Session = Depends(get_session)):
    """从持仓截图 OCR 识别并导入。使用 Claude Vision 识别。"""
    settings = get_settings()
    if not settings.anthropic_api_key:
        raise HTTPException(500, "OCR 功能需要配置 ANTHROPIC_API_KEY")

    content = await file.read()
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(400, "请上传图片文件（PNG/JPG）")

    b64_image = base64.standard_b64encode(content).decode("utf-8")
    media_type = file.content_type or "image/png"

    from anthropic import Anthropic
    client = Anthropic(api_key=settings.anthropic_api_key)

    response = client.messages.create(
        model=settings.anthropic_model,
        max_tokens=2000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media_type, "data": b64_image},
                },
                {
                    "type": "text",
                    "text": (
                        "请识别这张基金持仓截图中的所有基金信息。"
                        "输出 JSON 数组，每个元素包含：\n"
                        '{"fund_code": "6位代码", "fund_name": "名称", "shares": 份额数字, "cost_price": 成本净值数字}\n'
                        "如果某个字段看不清就填 null。只输出 JSON，不要其他文字。"
                    ),
                },
            ],
        }],
    )

    text = response.content[0].text.strip()
    # 提取 JSON
    try:
        if text.startswith("["):
            items = json.loads(text)
        else:
            start = text.index("[")
            end = text.rindex("]") + 1
            items = json.loads(text[start:end])
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(422, f"OCR 结果解析失败: {text[:200]}") from e

    imported = []
    for item in items:
        fund_code = item.get("fund_code")
        if not fund_code:
            continue
        holding = PortfolioHolding(
            fund_code=str(fund_code).strip(),
            fund_name=item.get("fund_name") or fund_code,
            shares=float(item.get("shares") or 0),
            cost_price=float(item.get("cost_price") or 1.0),
            buy_date=date.today(),
            category="equity",
        )
        db.add(holding)
        imported.append(item)

    db.commit()
    return {"status": "ok", "imported_count": len(imported), "items": imported, "raw_ocr": text}


# === 辅助函数 ===

def _find_value(row: dict, keys: list[str]) -> str | None:
    for k in keys:
        for col, val in row.items():
            if k in col:
                return val.strip() if val else None
    return None


def _find_float(row: dict, keys: list[str]) -> float | None:
    val = _find_value(row, keys)
    if val:
        try:
            return float(val.replace(",", "").replace("份", "").replace("元", ""))
        except ValueError:
            return None
    return None


def _parse_date(s: str) -> date:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            from datetime import datetime
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return date.today()


def _normalize_category(raw: str) -> str:
    raw = raw.lower().strip()
    if any(k in raw for k in ["股票", "权益", "equity", "指数"]):
        return "equity"
    elif any(k in raw for k in ["债", "bond"]):
        return "bond"
    elif any(k in raw for k in ["货币", "money"]):
        return "money"
    elif any(k in raw for k in ["混合", "hybrid"]):
        return "hybrid"
    return "equity"
